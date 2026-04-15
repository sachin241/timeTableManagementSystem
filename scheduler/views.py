# views.py — final integrated version
# Room-aware, teacher-limit-aware, scheduler-compatible views

import json
import os
from io import BytesIO

from django.conf import settings
from django.http import JsonResponse, HttpResponse, Http404
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from .models import Subject, Teacher, TimeSlot, Timetable, ClassSection, Room
from .parser import parse_and_store
from .scheduler import (
    generate_time_slots,
    persist_time_slots,
    generate_timetable,
    debug_schedule,
    partial_regenerate,
    SchedulingError,
    DEFAULT_CONFIG,
)

ALLOWED_EXTENSIONS = (".csv", ".xlsx", ".pdf", ".docx")
DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def _json_body(request):
    if request.content_type and "application/json" in request.content_type:
        try:
            return json.loads(request.body or b"{}")
        except json.JSONDecodeError:
            return {}
    return request.POST.dict()


def _parse_days(value):
    if not value:
        return []
    if isinstance(value, list):
        return [d.strip() for d in value if str(d).strip()]
    if isinstance(value, str):
        return [d.strip() for d in value.split(",") if d.strip()]
    return []


def _parse_config_from_body(body):
    config = {}

    for key in ("start_time", "end_time", "break_start", "break_end"):
        val = body.get(key, "")
        if isinstance(val, str) and val.strip():
            config[key] = val.strip()

    if body.get("slot_duration") not in (None, ""):
        try:
            config["slot_duration"] = int(body["slot_duration"])
        except (TypeError, ValueError):
            raise ValueError('"slot_duration" must be an integer.')

    days = _parse_days(body.get("days", ""))
    if days:
        config["days"] = days

    return config


def _slot_label(slot):
    return f"{slot.start_time.strftime('%H:%M')}–{slot.end_time.strftime('%H:%M')}"


def _build_conflict_list(time_slot, room=None, teacher=None, exclude_id=None):
    conflicts = []

    qs = Timetable.objects.select_related("subject", "teacher", "room", "section", "time_slot").filter(
        time_slot=time_slot
    )
    if exclude_id:
        qs = qs.exclude(pk=exclude_id)

    if teacher is not None:
        for entry in qs.filter(teacher=teacher):
            conflicts.append(
                f"{teacher.name} is already teaching {entry.subject.name} in [{entry.section.name}] at {time_slot}"
            )

    if room is not None:
        for entry in qs.filter(room=room):
            conflicts.append(
                f"Room {room.name} is already used for {entry.subject.name} in [{entry.section.name}] at {time_slot}"
            )

    return conflicts


def _teacher_load_payload(teacher):
    entries = Timetable.objects.filter(teacher=teacher)
    assigned_hours = entries.count()
    max_hours = teacher.max_hours_per_week or 0
    remaining = max(max_hours - assigned_hours, 0)

    section_breakdown = {}
    subject_breakdown = {}

    for entry in entries.select_related("section", "subject"):
        section_breakdown[entry.section.name] = section_breakdown.get(entry.section.name, 0) + 1
        subject_breakdown[entry.subject.name] = subject_breakdown.get(entry.subject.name, 0) + 1

    utilization = round((assigned_hours / max_hours) * 100, 1) if max_hours > 0 else 0.0

    return {
        "id": teacher.id,
        "name": teacher.name,
        "assigned_hours": assigned_hours,
        "max_hours_per_week": max_hours,
        "remaining_hours": remaining,
        "utilization": utilization,
        "overloaded": assigned_hours > max_hours if max_hours > 0 else False,
        "sections": section_breakdown,
        "subjects": subject_breakdown,
    }


# ---------------------------------------------------------------------
# Landing Page
# ---------------------------------------------------------------------

def index(request):
    ctx = {
        "default_start": DEFAULT_CONFIG["start_time"],
        "default_end": DEFAULT_CONFIG["end_time"],
        "default_duration": DEFAULT_CONFIG["slot_duration"],
        "default_break_start": DEFAULT_CONFIG["break_start"],
        "default_break_end": DEFAULT_CONFIG["break_end"],
        "default_days": ",".join(DEFAULT_CONFIG["days"]),
        "days_list": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
        "default_days_set": set(DEFAULT_CONFIG["days"]),
    }
    return render(request, "scheduler/index.html", ctx)


# ---------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def upload_file(request):
    if "csv_file" not in request.FILES:
        return JsonResponse({"success": False, "error": "No file uploaded."}, status=400)

    uploaded = request.FILES["csv_file"]
    ext = os.path.splitext(uploaded.name)[1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        return JsonResponse(
            {
                "success": False,
                "error": f'Unsupported format "{ext}". Supported: {", ".join(ALLOWED_EXTENSIONS)}',
            },
            status=400,
        )

    upload_dir = os.path.join(settings.MEDIA_ROOT, "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, f"input{ext}")

    with open(file_path, "wb+") as f:
        for chunk in uploaded.chunks():
            f.write(chunk)

    try:
        data = parse_and_store(file_path)
        request.session["upload_path"] = file_path

        total_hours = sum(data["subject_hours_map"].values())
        labs = [s for s in data["subjects"] if getattr(s, "subject_type", "theory") == "lab"]
        sections = [s.name for s in data["sections"]]

        return JsonResponse(
            {
                "success": True,
                "message": f"Parsed successfully ({ext.upper().strip('.') } format).",
                "subjects": len(data["subjects"]),
                "teachers": len(data["teachers"]),
                "rooms": len(data.get("rooms", [])),
                "total_hours": total_hours,
                "labs_count": len(labs),
                "sections": sections,
            }
        )
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Unexpected error: {str(e)}"}, status=500)


# ---------------------------------------------------------------------
# Generate
# ---------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def generate(request):
    upload_path = request.session.get("upload_path") or request.session.get("csv_path")
    if not upload_path or not os.path.exists(upload_path):
        return JsonResponse(
            {"success": False, "error": "No uploaded file found. Please upload a file first."},
            status=400,
        )

    body = _json_body(request)

    try:
        config = _parse_config_from_body(body)

        data = parse_and_store(upload_path)

        slot_dicts = generate_time_slots(config)
        if not slot_dicts:
            return JsonResponse(
                {"success": False, "error": "No time slots generated. Check config."},
                status=400,
            )

        orm_slots = persist_time_slots(slot_dicts)

        success = generate_timetable(data, orm_slots)

        if not success:
            return JsonResponse(
                {"success": False, "error": "Scheduler could not satisfy all constraints."},
                status=422,
            )

        count = Timetable.objects.count()
        sections = list(ClassSection.objects.values_list("name", flat=True))

        return JsonResponse(
            {
                "success": True,
                "message": f"Timetable generated: {count} sessions across {len(sections)} section(s).",
                "count": count,
                "total_slots": len(orm_slots),
                "sections": sections,
            }
        )

    except SchedulingError as e:
        return JsonResponse({"success": False, "error": str(e)}, status=422)
    except ValueError as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Unexpected error: {str(e)}"}, status=500)


# ---------------------------------------------------------------------
# Timetable page
# ---------------------------------------------------------------------

def timetable_view(request):
    has_data = Timetable.objects.exists()
    sections = list(ClassSection.objects.values_list("name", flat=True))
    return render(
        request,
        "scheduler/timetable.html",
        {
            "has_data": has_data,
            "sections": sections,
        },
    )


# ---------------------------------------------------------------------
# Timetable JSON
# ---------------------------------------------------------------------

def api_timetable(request):
    requested = request.GET.get("section", "").strip()
    all_sections = list(ClassSection.objects.values_list("name", flat=True))

    entries = Timetable.objects.select_related("subject", "teacher", "room", "time_slot", "section")

    if requested and requested in all_sections:
        entries = entries.filter(section__name=requested)
        current_section = requested
    elif all_sections:
        entries = entries.filter(section__name=all_sections[0])
        current_section = all_sections[0]
    else:
        current_section = ""

    if not entries.exists():
        return JsonResponse(
            {
                "days": [],
                "slots": [],
                "grid": {},
                "summary": {},
                "sections_list": all_sections,
                "current_section": current_section,
            }
        )

    slot_labels_set = set()
    grid = {}
    subject_count = {}

    colors = [
        "#6366f1", "#8b5cf6", "#ec4899", "#f59e0b",
        "#10b981", "#3b82f6", "#ef4444", "#14b8a6",
        "#f97316", "#84cc16",
    ]
    subject_colors = {}
    color_idx = 0

    for entry in entries:
        slot = entry.time_slot
        label = _slot_label(slot)
        slot_labels_set.add((slot.slot_index, label))

        if entry.subject.name not in subject_colors:
            subject_colors[entry.subject.name] = colors[color_idx % len(colors)]
            color_idx += 1

        grid.setdefault(slot.day, {})
        grid[slot.day][label] = {
            "subject": entry.subject.name,
            "teacher": entry.teacher.name,
            "room": entry.room.name if getattr(entry, "room", None) else "",
            "color": subject_colors[entry.subject.name],
            "is_lab": entry.subject.subject_type == "lab",
            "is_continuation": entry.is_lab_continuation,
            "entry_id": entry.id,
            "explanation": entry.explanation or {},
        }
        subject_count[entry.subject.name] = subject_count.get(entry.subject.name, 0) + 1

    sorted_labels = [label for _, label in sorted(slot_labels_set)]
    active_days = [d for d in DAY_ORDER if d in grid]

    return JsonResponse(
        {
            "days": active_days,
            "slots": sorted_labels,
            "grid": grid,
            "summary": {
                "total": entries.count(),
                "subjects": subject_count,
            },
            "sections_list": all_sections,
            "current_section": current_section,
        }
    )


# ---------------------------------------------------------------------
# Edit / Validate
# ---------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def api_edit(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON."}, status=400)

    entry_id = body.get("entry_id")
    new_subject_id = body.get("subject_id")
    new_teacher_id = body.get("teacher_id")
    new_room_id = body.get("room_id")

    if not entry_id:
        return JsonResponse({"success": False, "error": "entry_id is required."}, status=400)

    try:
        entry = Timetable.objects.select_related("time_slot", "section", "subject", "teacher", "room").get(pk=entry_id)
    except Timetable.DoesNotExist:
        return JsonResponse({"success": False, "error": "Entry not found."}, status=404)

    try:
        new_subject = Subject.objects.get(pk=new_subject_id) if new_subject_id else entry.subject
        new_teacher = Teacher.objects.get(pk=new_teacher_id) if new_teacher_id else entry.teacher
        new_room = Room.objects.get(pk=new_room_id) if new_room_id else entry.room
    except (Subject.DoesNotExist, Teacher.DoesNotExist, Room.DoesNotExist):
        return JsonResponse({"success": False, "error": "Subject, teacher, or room not found."}, status=404)

    if new_subject not in new_teacher.subjects.all():
        return JsonResponse(
            {"success": False, "error": f"{new_teacher.name} cannot teach {new_subject.name}."},
            status=409,
        )

    if new_subject.subject_type == "lab" and new_room and not new_room.is_lab:
        return JsonResponse(
            {"success": False, "error": f"Lab subject {new_subject.name} requires a lab room."},
            status=409,
        )

    conflicts = _check_conflicts(
        time_slot=entry.time_slot,
        section=entry.section,
        teacher=new_teacher,
        room=new_room,
        exclude_id=entry.id,
    )
    if conflicts:
        return JsonResponse(
            {
                "success": False,
                "error": "Conflict detected.",
                "conflicts": conflicts,
            },
            status=409,
        )

    entry.subject = new_subject
    entry.teacher = new_teacher
    if new_room is not None:
        entry.room = new_room

    entry.explanation = {
        "reason": f"Manually edited: updated to {new_subject.name}",
        "factors": [
            "User requested change",
            f"Teacher set to {new_teacher.name}",
            f"Room set to {new_room.name if new_room else 'unchanged'}",
        ],
        "sc_score": 0.0,
        "phase": "manual_edit",
    }
    entry.save()

    return JsonResponse(
        {
            "success": True,
            "message": f"Updated: {entry.time_slot} → {new_subject.name} ({new_teacher.name})",
        }
    )


@csrf_exempt
@require_http_methods(["POST"])
def api_validate(request):
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON."}, status=400)

    entry_id = body.get("entry_id")
    new_teacher_id = body.get("teacher_id")
    new_room_id = body.get("room_id")

    if not entry_id:
        return JsonResponse({"success": False, "error": "entry_id required."}, status=400)

    try:
        entry = Timetable.objects.select_related("time_slot", "section", "subject", "teacher", "room").get(pk=entry_id)
    except Timetable.DoesNotExist:
        return JsonResponse({"success": False, "error": "Entry not found."}, status=404)

    new_teacher = entry.teacher
    new_room = entry.room

    if new_teacher_id:
        try:
            new_teacher = Teacher.objects.get(pk=new_teacher_id)
        except Teacher.DoesNotExist:
            return JsonResponse({"success": False, "error": "Teacher not found."}, status=404)

    if new_room_id:
        try:
            new_room = Room.objects.get(pk=new_room_id)
        except Room.DoesNotExist:
            return JsonResponse({"success": False, "error": "Room not found."}, status=404)

    if entry.subject not in new_teacher.subjects.all():
        return JsonResponse(
            {
                "valid": False,
                "conflicts": [f"{new_teacher.name} cannot teach {entry.subject.name}."],
            }
        )

    if entry.subject.subject_type == "lab" and new_room and not new_room.is_lab:
        return JsonResponse(
            {
                "valid": False,
                "conflicts": [f"Lab subject {entry.subject.name} requires a lab room."],
            }
        )

    conflicts = _check_conflicts(
        time_slot=entry.time_slot,
        section=entry.section,
        teacher=new_teacher,
        room=new_room,
        exclude_id=entry.id,
    )

    return JsonResponse(
        {
            "valid": len(conflicts) == 0,
            "conflicts": conflicts,
        }
    )


def _check_conflicts(time_slot, section, teacher, room=None, exclude_id=None):
    conflicts = []

    teacher_qs = Timetable.objects.filter(time_slot=time_slot, teacher=teacher)
    if exclude_id:
        teacher_qs = teacher_qs.exclude(pk=exclude_id)

    for conflict_entry in teacher_qs:
        conflicts.append(
            f"{teacher.name} is already teaching {conflict_entry.subject.name} in [{conflict_entry.section.name}] at {time_slot}"
        )

    if room is not None:
        room_qs = Timetable.objects.filter(time_slot=time_slot, room=room)
        if exclude_id:
            room_qs = room_qs.exclude(pk=exclude_id)

        for conflict_entry in room_qs:
            conflicts.append(
                f"{room.name} is already used for {conflict_entry.subject.name} in [{conflict_entry.section.name}] at {time_slot}"
            )

    return conflicts


# ---------------------------------------------------------------------
# Explain
# ---------------------------------------------------------------------

def api_explain(request, section, day, slot):
    """
    GET /api/explain/<section>/<day>/<slot>/
    slot can be either a label like 09:00–10:00 or a slot index string.
    """
    section = section.strip()
    day = day.strip()
    slot_raw = slot.strip().replace("-", "–")

    entries = Timetable.objects.select_related("subject", "teacher", "room", "time_slot", "section").filter(
        section__name=section,
        time_slot__day=day,
    )

    if not entries.exists():
        return JsonResponse({"found": False, "explanation": {}})

    # Match by label first, then by slot index if numeric
    for e in entries:
        label = _slot_label(e.time_slot)
        if label == slot_raw or label.replace("–", "-") == slot_raw.replace("–", "-"):
            return JsonResponse(
                {
                    "found": True,
                    "subject": e.subject.name,
                    "teacher": e.teacher.name,
                    "room": e.room.name if getattr(e, "room", None) else "",
                    "slot": label,
                    "day": day,
                    "section": section,
                    "explanation": e.explanation or {},
                    "is_lab": e.subject.subject_type == "lab",
                    "is_continuation": e.is_lab_continuation,
                }
            )

    if slot_raw.isdigit():
        for e in entries:
            if str(e.time_slot.slot_index) == slot_raw:
                return JsonResponse(
                    {
                        "found": True,
                        "subject": e.subject.name,
                        "teacher": e.teacher.name,
                        "room": e.room.name if getattr(e, "room", None) else "",
                        "slot": _slot_label(e.time_slot),
                        "day": day,
                        "section": section,
                        "explanation": e.explanation or {},
                        "is_lab": e.subject.subject_type == "lab",
                        "is_continuation": e.is_lab_continuation,
                    }
                )

    return JsonResponse({"found": False, "explanation": {}})


# ---------------------------------------------------------------------
# Debug schedule
# ---------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def api_debug_schedule(request):
    upload_path = request.session.get("upload_path") or request.session.get("csv_path")
    if not upload_path or not os.path.exists(upload_path):
        return JsonResponse({"success": False, "error": "No uploaded file. Upload first."}, status=400)

    body = _json_body(request)
    try:
        config = _parse_config_from_body(body)
        data = parse_and_store(upload_path)
        slot_dicts = generate_time_slots(config)
        orm_slots = persist_time_slots(slot_dicts)
        reports = debug_schedule(data, orm_slots)

        return JsonResponse({"success": True, "reports": reports})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


# ---------------------------------------------------------------------
# Teacher load
# ---------------------------------------------------------------------

def api_teacher_load(request):
    teachers = Teacher.objects.all()

    result = [_teacher_load_payload(teacher) for teacher in teachers]
    result.sort(key=lambda x: (-int(x["overloaded"]), -x["utilization"], x["name"]))

    return JsonResponse(
        {
            "teachers": result,
            "total_slots": TimeSlot.objects.count(),
            "total_sections": ClassSection.objects.count(),
        }
    )


def dashboard_view(request):
    return render(request, "scheduler/dashboard.html")


# ---------------------------------------------------------------------
# Partial regeneration
# ---------------------------------------------------------------------

@csrf_exempt
@require_http_methods(["POST"])
def api_partial_regen(request):
    """
    If only subject/teacher changes are requested, delegates to scheduler.partial_regenerate.
    If room_id is also supplied, updates the room after conflict checks.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON."}, status=400)

    entry_id = body.get("entry_id")
    if not entry_id:
        return JsonResponse({"success": False, "error": "entry_id required."}, status=400)

    room_id = body.get("room_id")

    try:
        result = partial_regenerate(
            entry_id=int(entry_id),
            new_subject_id=body.get("subject_id"),
            new_teacher_id=body.get("teacher_id"),
        )
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)

    # Optional room update after partial regenerate
    if room_id:
        try:
            entry = Timetable.objects.select_related("time_slot", "section", "subject", "teacher", "room").get(pk=int(entry_id))
            new_room = Room.objects.get(pk=room_id)

            room_conflicts = _check_conflicts(
                time_slot=entry.time_slot,
                section=entry.section,
                teacher=entry.teacher,
                room=new_room,
                exclude_id=entry.id,
            )
            if room_conflicts:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "Room conflict detected during partial regeneration.",
                        "conflicts": room_conflicts,
                    },
                    status=409,
                )

            if entry.subject.subject_type == "lab" and not new_room.is_lab:
                return JsonResponse(
                    {
                        "success": False,
                        "error": "Lab subject requires a lab room.",
                    },
                    status=409,
                )

            entry.room = new_room
            entry.save()
            result.setdefault("changed", []).append(
                f"[{entry.section.name}] {entry.time_slot} → room changed to {new_room.name}"
            )
            result["success"] = True
        except Room.DoesNotExist:
            return JsonResponse({"success": False, "error": "Room not found."}, status=404)
        except Timetable.DoesNotExist:
            return JsonResponse({"success": False, "error": "Entry not found."}, status=404)

    status_code = 200 if result.get("success") else 422
    return JsonResponse(result, status=status_code)


# ---------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------

def download_pdf(request):
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    section_name = request.GET.get("section", "")
    entries = Timetable.objects.select_related("subject", "teacher", "room", "time_slot", "section")

    if section_name:
        entries = entries.filter(section__name=section_name)

    if not entries.exists():
        raise Http404("No timetable data found.")

    sections_data = {}
    for entry in entries:
        sections_data.setdefault(entry.section.name, []).append(entry)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=20)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, alignment=TA_CENTER)

    story = []

    for sec_name, sec_entries in sorted(sections_data.items()):
        slot_set = {}
        grid = {}

        for e in sec_entries:
            slot = e.time_slot
            label = _slot_label(slot)
            slot_set[slot.slot_index] = label

            room_txt = f"\n[{e.room.name}]" if getattr(e, "room", None) else ""
            text = f"{e.subject.name}\n({e.teacher.name}){room_txt}"
            if e.subject.subject_type == "lab":
                text = f"🔬 {text}"

            grid.setdefault(slot.day, {})[label] = text

        sorted_labels = [v for _, v in sorted(slot_set.items())]
        active_days = [d for d in DAY_ORDER if d in grid]

        header = ["Time Slot"] + active_days
        table_data = [header]

        for sl in sorted_labels:
            row = [sl]
            for day in active_days:
                cell_text = grid.get(day, {}).get(sl, "—")
                row.append(Paragraph(cell_text.replace("\n", "<br/>"), cell_style))
            table_data.append(row)

        col_widths = [3 * cm] + [3.5 * cm] * len(active_days)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#4f46e5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.grey),
                    ("BACKGROUND", (0, 1), (0, -1), rl_colors.HexColor("#e0e7ff")),
                    ("ROWBACKGROUNDS", (1, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f5f3ff")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        story.append(Paragraph(f"Timetable — {sec_name}", title_style))
        story.append(Spacer(1, 0.5 * cm))
        story.append(table)
        story.append(Spacer(1, 1.5 * cm))

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="timetable.pdf"'
    return response


# ---------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------

def download_xlsx(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    section_name = request.GET.get("section", "all")
    entries = Timetable.objects.select_related("subject", "teacher", "room", "time_slot", "section")

    if section_name and section_name != "all":
        entries = entries.filter(section__name=section_name)

    if not entries.exists():
        raise Http404("No timetable data found.")

    sections_data = {}
    for entry in entries:
        sections_data.setdefault(entry.section.name, []).append(entry)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sec_name in sorted(sections_data.keys()):
        sec_entries = sections_data[sec_name]
        ws = wb.create_sheet(title=sec_name[:31])

        slot_set = {}
        grid = {}

        for e in sec_entries:
            slot = e.time_slot
            label = _slot_label(slot)
            slot_set[slot.slot_index] = label
            tag = " [LAB]" if e.subject.subject_type == "lab" else ""
            room_txt = f" [{e.room.name}]" if getattr(e, "room", None) else ""
            grid.setdefault(slot.day, {})[label] = f"{e.subject.name}{tag}{room_txt}\n{e.teacher.name}"

        sorted_labels = [v for _, v in sorted(slot_set.items())]
        active_days = [d for d in DAY_ORDER if d in grid]

        headers = ["Time Slot"] + active_days
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = cell_align
            cell.border = thin_border

        for row_idx, sl in enumerate(sorted_labels, 2):
            cell = ws.cell(row=row_idx, column=1, value=sl)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(bold=True)

            for col_idx, day in enumerate(active_days, 2):
                val = grid.get(day, {}).get(sl, "—")
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.alignment = cell_align
                c.border = thin_border

        ws.column_dimensions["A"].width = 16
        for i in range(2, len(active_days) + 2):
            col_letter = chr(64 + i)
            ws.column_dimensions[col_letter].width = 24

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="timetable.xlsx"'
    return response